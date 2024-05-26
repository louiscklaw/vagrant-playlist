#!/usr/bin/env python

import os
import sys
import re
from pprint import pprint

import json
from jsonpath_ng import jsonpath, parse

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image

import re


def LookupTeamPosition(team_to_lookup, manifest_json):
    team_rank = -99
    num_of_team = -99

    standings = manifest_json['tournament/standings.json']['info']
    num_of_team = len(standings)

    for standing in standings:
        if standing['teamName'] == team_to_lookup:
            team_rank = int(standing['teamRank'])
    return [team_rank, num_of_team]


def writeExcel(json_manifest, image_paths, report_filename):
    pattern = r"(?<={jp:)(.*)(?=})"

    excel_template_path = '/report_template.xlsx'

    os_cwd = os.getcwd()
    utils_path = os_cwd+'/utils'
    output_path = os_cwd + "/_output"
    images_path = os_cwd + "/_images"

    # Find and replace values
    find_value = 'Mainz'
    replace_value = '11111111'

    # Create a new workbook
    workbook = load_workbook(utils_path+excel_template_path)

    # Select the active sheet (first sheet by default)
    sheet = workbook.active

    for row in sheet.iter_rows():
        for cell in row:
            # print(cell.coordinate)

            if type(cell.value) == type('string'):
                result = re.search(pattern, cell.value)
                if result:
                    try:
                        jsonpath_expression = parse(result.group(0))
                        alignment = Alignment(
                            horizontal='center', vertical='center')
                        cell.alignment = alignment
                        cell.value = jsonpath_expression.find(json_manifest)[
                            0].value

                    except Exception as e:
                        pprint(e)
                        pprint("error:" + cell.value)
                        alignment = Alignment(
                            horizontal='center', vertical='center')
                        cell.alignment = alignment
                        cell.value = '---'

    # image_path = images_path+"/helloworld.jpg"

    current_cell = sheet.cell(row=1, column=9)
    for image_path in image_paths:
        print(image_path)
        image = Image(image_path)
        sheet.add_image(image, current_cell.coordinate)
        current_cell = current_cell.offset(row=0, column=17)

    workbook.save(output_path+'/'+report_filename)



def helloworld():
    print(os.getcwd())
    print("helloworld")
